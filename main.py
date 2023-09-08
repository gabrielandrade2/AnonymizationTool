import argparse
import csv
import os.path

import spacy
from openpyxl import load_workbook

ANONYMIZED_TAG = '[ANON]'

anonymization_count = {
    'Person': 0,
    'Location': 0,
    'Other': 0,
    'Stop-word tokens': 0,
    'Forced anonymization': 0,
}

nlp = spacy.load("ja_ginza")

force_anonymize_columns = []
force_anonymize_tokens = []
#Default stop words
stop_words = ['病院', 'クリニック', 'Dr']
out_dir = None

# Load name list csv
names_list = []
with open("names_list.csv") as csv_file:
    csv_reader = csv.reader(csv_file)

    # Skip the header row if it exists.
    header = next(csv_reader, None)

    # Iterate through each row in the CSV file and append it to the list.
    for row in csv_reader:
        names_list.append(row[0])
# Filter out single character names
names_list = set(filter(lambda x: len(x) > 1, names_list))

def extract_longest_sequence(tokens, target):
    all_sequences = []
    current_sequence = []

    for i, tok in enumerate(tokens):
        num = tok.i
        if i == 0 or num == tokens[i - 1].i + 1:
            current_sequence.append(tok)
        else:
            if current_sequence:
                all_sequences.append(current_sequence)
            current_sequence = [tok]

    if current_sequence:
        all_sequences.append(current_sequence)

    for sequence in all_sequences:
        if target in [tok.i for tok in sequence]:
            return sequence


def process_file(file):
    filename = os.path.split(file)[1]
    f = load_workbook(file)
    for sheet in f.worksheets:
        for column in sheet.iter_cols():
            if column[0].value in force_anonymize_columns:
                for cell in column[1:]:
                    cell.value = force_deidentify(str(cell.value))
                continue
            for cell in column[1:]:
                if isinstance(cell.value, str):
                    print(cell)
                    cell.value = deidentify(cell.value)

    out = os.path.join(out_dir, filename)
    if os.path.exists(out):
        out = out.replace(".xlsx", "_anon.xlsx")
    f.save(out)


def should_deidentify(token):
    if token.text in names_list:
        anonymization_count['Person'] += 1
        return True

    tags = token.tag_.split('-')
    # Remove 固有名詞, but only the ones detected as person name (人名), location(地名)
    # 一般 is controversial, can be used to remove company names, but also can remove medication names
    if '固有名詞' in tags:
        if '人名' in tags:
            anonymization_count['Person'] += 1
            return True
        elif '地名' in tags:
            anonymization_count['Location'] += 1
            return True
        # elif '一般' in tags:
        # TODO: Test if these can be removed safely in pair with stop-words

        #     anonymization_count['Other'] += 1
        #     return True

    # Remove all proper nouns detected as such
    if token.pos_ == 'PROPN' and '一般' not in tags:
        anonymization_count['Other'] += 1
        return True

    elif token.text in force_anonymize_tokens:
        anonymization_count["Force Anonymize Tokens"] += 1
        return True
    return False


def deidentify(text: str):
    """
    Method that performs the actual anonymization of texts. Can be called directly from other scripts in order to
    execute the anonymization logic in a single string.

    :param text: The text to be anonymized.
    :return: The anonymized text.
    """
    parsed = nlp(text)

    # Debug print
    # for sent in parsed.sents:
    #     for token in sent:
    #         print(
    #             token.i,
    #             token.orth_,
    #             token.pos_,
    #             token.tag_,
    #             token.dep_,
    #             token.head.i,
    #         )
    #     print('EOS')

    tokens = []
    for sent in parsed.sents:
        for token in sent:
            tokens.append(token)

    it = iter(enumerate(tokens))
    anonymized_text = []
    for i, token in it:
        anon = False

        if should_deidentify(token):
            if len(anonymized_text) == 0 or anonymized_text[-1] != ANONYMIZED_TAG:
                anonymized_text.append(ANONYMIZED_TAG)
            anon = True

        # Check for compound nouns. If some part of the compound noun is a proper noun, remove it
        compound_end = -1
        if token.dep_ == 'compound' and (token.pos_ == 'NOUN' or token.pos_ == 'PROPN' or '名詞' in token.head.tag_):
            head = token.head
            children = list(filter(lambda x: x.dep_ == 'compound', list(head.children)))
            neighbors = extract_longest_sequence(children, token.i)
            neighbors = list(filter(lambda x: x.i != token.i, neighbors))
            if neighbors:
                compound_end = neighbors[-1].i
                for neighbor in neighbors:
                    if should_deidentify(neighbor):
                        if len(anonymized_text) == 0 or anonymized_text[-1] != ANONYMIZED_TAG:
                            anonymized_text.append(ANONYMIZED_TAG)
                        anon = True
                        break


        # if next token (or after the compound) is a stop word,
        if compound_end != -1 and len(tokens) > (compound_end + 1) and tokens[compound_end + 1].text in stop_words:
            to_anon = False
            for j in range(i, compound_end + 1):
                if tokens[j].pos_ == 'PROPN' or '名詞' in tokens[j].tag_ or tokens[j].pos_ == 'NOUN':
                    to_anon = True
            if to_anon:
                anonymization_count["Stop-word tokens"] += 1
                anonymized_text.append(ANONYMIZED_TAG)
                anon = True
                for j in range(i, compound_end + 1):
                    i, token = next(it)
                anonymized_text.append(token.text)

        elif len(tokens) > (i + 1) and tokens[i + 1].text in stop_words:
            if token.pos_ == 'PROPN' or '名詞' in token.tag_ or token.pos_ == 'NOUN':
                anonymization_count["Stop-word tokens"] += 1
                anonymized_text.append(ANONYMIZED_TAG)
                anon = True
                i, token = next(it)
                anonymized_text.append(token.text)

        if not anon:
            anonymized_text.append(token.text)
    return "".join(anonymized_text)


def force_deidentify(text):
    for _ in text.split(" "):
        anonymization_count["Forced anonymization"] += 1
    return " ".join([ANONYMIZED_TAG for _ in text.split(" ")])


def process_directory(directory):
    files = os.listdir(directory)
    filtered = filter(lambda x: x.endswith(".xlsx") and not x.startswith("."), files)
    return [os.path.join(directory, f) for f in filtered]


def main(input: str, output: str, force_anonymize_columns: list = None, force_anonymize_tokens: list = None,
         stop_words: list = None):
    """
    Main function for anonymizing Excel files, called when executing this script directly.

    :param input: The input file(s) or directory(ies)
    :param output: The output directory
    :param force_anonymize_columns: Columns to be forcibly anonymized, regardless of the content type
    :param force_anonymize_tokens: Special tokens that should always be anonymized
    :param stop_words: Special words that implicate the previous word should be anonymized, e.g. "病院" or "クリニック"
    :return:
    """
    global out_dir
    globals()['force_anonymize_columns'] = force_anonymize_columns
    globals()['force_anonymize_tokens'] = force_anonymize_tokens
    globals()['stop_words'] = stop_words
    out_dir = output

    # Parse file list
    files = []
    for path in input:
        if os.path.isfile(path) and path.endswith(".xlsx"):
            files.append(path)
        elif os.path.isdir(path):
            files.extend(process_directory(path))
        else:
            print(f"Invalid file: {path}")

    os.makedirs(out_dir, exist_ok=True)

    for file in files:
        try:
            process_file(file)
        except Exception as e:
            print(f"Error processing {file}: {e}")
            continue

    return anonymization_count, len(files)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Excel file anonymizer tool')
    parser.add_argument('--input', type=str, nargs='+', required=True,
                        help='Input file(s) or directory(ies) path')
    parser.add_argument('--output', type=str, required=True, help='Anonymized output folder')
    parser.add_argument('--force_anonymize_columns', type=str, nargs='+',
                        help='Columns to be forcibly anonymized, regardless of the content type')
    parser.add_argument('--force_anonymize_tokens', type=str, nargs='+',
                        help='Special tokens that should always be anonymized')
    parser.add_argument('--stop_words', type=str, nargs='+',
                        help='''Special words that implicate the previous word should be anonymized, e.g. "病院" or "クリニック" \n
                        Default: {}'''.format(stop_words))

    args = parser.parse_args()

    count, files = main(args.input, args.output, args.force_anonymize_columns, args.force_anonymize_tokens, args.stop_words)
    print("Processed files:", files)
    print("Anonymized tokens:", count)

