import argparse
import glob
import os.path
import re

import MeCab
from openpyxl import load_workbook
from unidic import unidic

ANONYMIZED_TAG = '[ANON]'

anonymization_count = {
    'Person': 0,
    'Location': 0,
    'Organization': 0,
    'Other': 0,
}
tagger = MeCab.Tagger('-d "{}"'.format(unidic.DICDIR))


def should_deidentify(token):
    # Remove all proper nouns
    if (token[1][0] == '名詞' and token[1][1] == '固有名詞'):
        # Person name
        if token[1][2] == '人名':
            anonymization_count['Person'] += 1
        # Location
        elif token[1][2] == '地名':
            anonymization_count['Location'] += 1
        # Organization
        elif token[1][2] == '一般' or token[1][3] == '一般' or token[1][2] == "地域":
            anonymization_count['Organization'] += 1
        else:
            anonymization_count["Other"] += 1
        return True
    elif token[0] in ["高原", "えきさいかい"]:
        return True
    return False

def get_mecab_parsing(text):
    return [[chunk.split('\t')[0], tuple(chunk.split('\t')[1].split(','))] for chunk in
              tagger.parse(text).splitlines()[:-1]]

def deidentify(text):
    parsed = get_mecab_parsing(text)
    for i, token in enumerate(parsed):
        if should_deidentify(token):
            text = text.replace(token[0], ANONYMIZED_TAG)
        if token[0] in ["病院", "クリニック"]:
            text = text.replace(parsed[i-1][0], ANONYMIZED_TAG)
    return text

def force_deidentify(text):
    return " " .join([ANONYMIZED_TAG for _ in text.split(" ")])

if __name__ == '__main__':
    # parser = argparse.ArgumentParser(description='Anonymize Nagoya University Hospital records')
    # parser.add_argument('--input', type=str, help='Input file path')
    # parser.add_argument('--output', type=str, help='Anonymized output file path')
    # args = parser.parse_args()
    #

    force_remove = ["患者番号", "発生者氏名", "更新者氏名", "患者ID", "患者ID"]

    directory ="C:\\Users\\精神科医局\\Desktop\\カルテデータ\\~2018\\"
    files = ["医師記録.xlsx", "看護記録.xlsx"]

    for filename in files:
        file = os.path.join(directory, filename)
        f = load_workbook(file)
        for sheet in f.worksheets:
            for column in sheet.iter_cols():
                if column[0].value in force_remove:
                    for cell in column[1:]:
                        cell.value = force_deidentify(str(cell.value))
                    continue
                for cell in column[1:]:
                    if isinstance(cell.value, str):
                        cell.value = deidentify(cell.value)

        out_dir = directory.replace("カルテデータ", "deidentified")
        os.makedirs(out_dir,exist_ok=True)
        f.save(os.path.join(out_dir, filename))

    # print(deidentify("田中と申します。大阪にすんでいます。"))
    print(anonymization_count)
